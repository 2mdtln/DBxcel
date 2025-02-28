import sqlite3
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
import json
import os
import webbrowser
import requests
import subprocess
import sys

def install_libraries():
    libraries = ["openpyxl", "requests"]
    for lib in libraries:
        try:
            __import__(lib)
        except ImportError:
            print(f"{lib} kütüphanesi eksik. Yükleniyor...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", lib])
            print(f"{lib} başarıyla yüklendi!")

install_libraries()

class ExportApp:
    def __init__(self, root):
        self.root = root
        self.root.withdraw()
        if self.check_for_updates():
            self.root.destroy()
            return
        self.root.deiconify()

        self.root.title("Veritabanını Excel'e Aktar")
        self.load_last_used_params()
        self.last_db_label = tk.Label(root, text=f"Son kullanılan DB: {self.last_db if self.last_db else 'Yok'}")
        self.last_db_label.pack()

        self.last_table_label = tk.Label(root,
                                         text=f"Son kullanılan Tablo: {self.last_table if self.last_table else 'Yok'}")
        self.last_table_label.pack()

        self.table_entry = tk.Entry(root)

        def clear_entry(event):
            if event.widget.get() == "Tablo adını girin":
                event.widget.delete(0, tk.END)

        self.table_entry.bind("<FocusIn>", clear_entry)

        self.table_entry.insert(0, self.last_table if self.last_table else "Tablo adını girin")
        self.table_entry.pack()

        button_frame = tk.Frame(root)
        button_frame.pack()

        self.db_file_button = tk.Button(button_frame, text="DB Dosyasını Seç", command=self.select_db_file)
        self.db_file_button.pack(side=tk.LEFT, padx=10)

        self.export_button = tk.Button(button_frame, text="Excel'e Aktar", command=self.export_to_excel)
        self.export_button.pack(side=tk.LEFT, padx=10)

        self.credit_label = tk.Label(root, text="ASAL Kütüphanesi için yapılmıştır - Şubat 2025", fg="black",
                                     font=("Helvetica", 10, "bold"))
        self.credit_label.pack()

        github_frame = tk.Frame(root)
        github_frame.pack()

        self.github_credit_1 = tk.Label(github_frame, text="@2mdtln", fg="blue", cursor="hand2")
        self.github_credit_1.pack(side=tk.LEFT, padx=5)
        self.github_credit_1.bind("<Button-1>", lambda e: webbrowser.open("https://github.com/2mdtln"))

        self.github_credit_2 = tk.Label(github_frame, text="@Restilov", fg="blue", cursor="hand2")
        self.github_credit_2.pack(side=tk.LEFT, padx=5)
        self.github_credit_2.bind("<Button-1>", lambda e: webbrowser.open("https://github.com/Restilov"))

        self.version_label = tk.Label(root, text="v1.0.1", fg="black", font=("Helvetica", 8))
        self.version_label.pack()

        self.db_file = self.last_db
        self.table_name = self.last_table

        self.root.geometry("350x150")
        self.root.resizable(False, False)

        # Tabloya tıklandığında kaybolması için event ekleme
        self.root.bind("<Button-1>", self.hide_table)

    def hide_table(self, event):
        self.last_table_label.pack_forget()

    def load_last_used_params(self):
        if os.path.exists("params.json"):
            try:
                with open("params.json", "r") as f:
                    params = json.load(f)
                    self.last_db = params.get("last_db", None)
                    self.last_table = params.get("last_table", None)
            except Exception as e:
                print(f"Son kullanılan parametreleri yüklerken hata oluştu: {e}")
                self.last_db = None
                self.last_table = None
        else:
            self.last_db = None
            self.last_table = None

    def save_last_used_params(self):
        params = {
            "last_db": self.db_file,
            "last_table": self.table_name
        }
        try:
            with open("params.json", "w") as f:
                json.dump(params, f)
        except Exception as e:
            print(f"Son kullanılan parametreleri kaydederken hata oluştu: {e}")

    def select_db_file(self):
        file = filedialog.askopenfilename(filetypes=[["SQLite DB dosyaları", "*.db;*.db3"]])
        if file:
            self.db_file = file
            self.last_db_label.config(text=f"Son kullanılan DB: {file}")
            self.root.title(f"Veri aktarılıyor: {file}")
            self.save_last_used_params()

    def export_to_excel(self):
        self.table_name = self.table_entry.get()
        if not self.db_file or self.table_name == "Tablo adını girin":
            messagebox.showerror("Sen de benim hatalarımdan birisin",
                                 "Lütfen bir veritabanı seçin ve tablo adını girin.")
            return

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[["Excel Dosyası", "*.xlsx"]])

        if not save_path:
            return
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()

            cursor.execute(f"SELECT * FROM {self.table_name}")
            rows = cursor.fetchall()
            columns = [description[0] for description in cursor.description]

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            for col_num, column in enumerate(columns, 1):
                ws.cell(row=1, column=col_num, value=column)

            for row_num, row in enumerate(rows, 2):
                for col_num, value in enumerate(row, 1):
                    if isinstance(value, bytes):
                        try:
                            value = value.decode('utf-8')
                        except UnicodeDecodeError:
                            value = "<Binary Veri>"
                    ws.cell(row=row_num, column=col_num, value=value)

            wb.save(save_path)
            messagebox.showinfo("Başarılı", f"Veriler '{save_path}' dosyasına başarıyla aktarıldı.")

            self.save_last_used_params()

        except Exception as e:
            messagebox.showerror("Sen de benim hatalarımdan birisin", f"Bir hata oluştu: {e}")
        finally:
            conn.close()

    def check_for_updates(self):
        current_version = "v1.0.2"
        repo_url = "https://api.github.com/repos/2mdtln/DBxcel/releases/latest"
        update_found = False
        try:
            response = requests.get(repo_url)
            if response.status_code == 200:
                latest_release = response.json()
                latest_version = latest_release['tag_name']

                if latest_version != current_version:
                    update_msg = f"Yeni bir sürüm mevcut: {latest_version}. Güncellemek ister misiniz?"
                    if messagebox.askyesno("Güncelleme Bulundu", update_msg):
                        webbrowser.open(latest_release['html_url'])
                    update_found = True
        except Exception as e:
            print(f"Update check failed: {e}")

        return update_found

if __name__ == "__main__":
    root = tk.Tk()
    app = ExportApp(root)
    root.mainloop()
