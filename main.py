import sqlite3
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
import json
import os
import webbrowser


class ExportApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Veritabanını Excel'e Aktar")
        self.load_last_used_params()
        self.last_db_label = tk.Label(root, text=f"Son kullanılan DB: {self.last_db if self.last_db else 'Yok'}")
        self.last_db_label.pack()

        self.last_table_label = tk.Label(root, text=f"Son kullanılan Tablo: {self.last_table if self.last_table else 'Yok'}")
        self.last_table_label.pack()

        self.table_entry = tk.Entry(root)
        self.table_entry.insert(0, self.last_table if self.last_table else "Tablo adını girin")
        self.table_entry.pack()

        button_frame = tk.Frame(root)
        button_frame.pack()

        self.db_file_button = tk.Button(button_frame, text="DB Dosyasını Seç", command=self.select_db_file)
        self.db_file_button.pack(side=tk.LEFT, padx=10)

        self.export_button = tk.Button(button_frame, text="Excel'e Aktar", command=self.export_to_excel)
        self.export_button.pack(side=tk.LEFT, padx=10)

        self.github_credit_1 = tk.Label(root, text="@2mdtln", fg="blue", cursor="hand2")
        self.github_credit_1.pack()
        self.github_credit_1.bind("<Button-1>", lambda e: webbrowser.open("https://github.com/2mdtln"))

        self.github_credit_2 = tk.Label(root, text="@Restilov", fg="blue", cursor="hand2")
        self.github_credit_2.pack()
        self.github_credit_2.bind("<Button-1>", lambda e: webbrowser.open("https://github.com/Restilov"))

        self.credit_label = tk.Label(root, text="Asal için yapılmıştır", fg="black", font=("Helvetica", 10, "bold"))
        self.credit_label.pack()

        self.db_file = self.last_db
        self.table_name = self.last_table

        self.root.geometry("350x220")

    def load_last_used_params(self):
        if os.path.exists("params.json"):
            try:
                with open("params.json", "r") as f:
                    params = json.load(f)
                    self.last_db = params.get("last_db", None)
                    self.last_table = params.get("last_table", None)
            except Exception as e:
                print(f"Son kullanılan parametreleri yüklerken hata oluştu: {e}")
                self.last_db = None
                self.last_table = None
        else:
            self.last_db = None
            self.last_table = None

    def save_last_used_params(self):
        params = {
            "last_db": self.db_file,
            "last_table": self.table_name
        }
        try:
            with open("params.json", "w") as f:
                json.dump(params, f)
        except Exception as e:
            print(f"Son kullanılan parametreleri kaydederken hata oluştu: {e}")

    def select_db_file(self):
        file = filedialog.askopenfilename(filetypes=[("SQLite DB dosyaları", "*.db;*.db3")])
        if file:
            self.db_file = file
            self.last_db_label.config(text=f"Son kullanılan DB: {file}")
            self.root.title(f"Veri aktarılıyor: {file}")
            self.save_last_used_params()

    def export_to_excel(self):
        self.table_name = self.table_entry.get()

        if not self.db_file or self.table_name == "Tablo adını girin":
            messagebox.showerror("Hata", "Lütfen bir veritabanı seçin ve tablo adını girin.")
            return

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Dosyası", "*.xlsx")])
        if not save_path:
            return

        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()

            cursor.execute(f"SELECT * FROM {self.table_name}")
            rows = cursor.fetchall()
            columns = [description[0] for description in cursor.description]

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            for col_num, column in enumerate(columns, 1):
                ws.cell(row=1, column=col_num, value=column)

            for row_num, row in enumerate(rows, 2):
                for col_num, value in enumerate(row, 1):
                    if isinstance(value, bytes):
                        try:
                            value = value.decode('utf-8')
                        except UnicodeDecodeError:
                            value = "<Binary Veri>"
                    ws.cell(row=row_num, column=col_num, value=value)

            wb.save(save_path)
            messagebox.showinfo("Başarılı", f"Veriler '{save_path}' dosyasına başarıyla aktarıldı.")
            
            self.save_last_used_params()

        except Exception as e:
            messagebox.showerror("Hata", f"Bir hata oluştu: {e}")
        finally:
            conn.close()

if __name__ == "__main__":
    root = tk.Tk()
    app = ExportApp(root)
    root.mainloop()
